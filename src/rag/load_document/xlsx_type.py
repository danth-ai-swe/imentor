from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import List, Tuple

import openpyxl

from src.rag.load_document.base_document import BaseDocumentExtractor
from src.utils.logger_utils import log


class XLSXExtractor(BaseDocumentExtractor):
    def _extract_blocks(self, file_path: Path) -> List[Tuple[int, str, bytes | str]]:
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        all_blocks: List[Tuple[int, str, bytes | str]] = []
        sheet_names = wb.sheetnames

        def _process_sheet(
                sheet_idx: int, sheet_name: str
        ) -> List[Tuple[int, str, bytes | str]]:
            ws = wb[sheet_name]
            rows_text: List[str] = []
            for row in ws.iter_rows(values_only=True):
                # Filter out fully empty rows
                row_values = [
                    str(cell).strip() if cell is not None else ""
                    for cell in row
                ]
                if any(row_values):
                    rows_text.append(" | ".join(row_values))
            if not rows_text:
                return []
            sheet_text = f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows_text)
            return [(sheet_idx * 100_000, "text", sheet_text)]

        with ThreadPoolExecutor(max_workers=self.max_workers) as pool:
            futures = {
                pool.submit(_process_sheet, idx, name): idx
                for idx, name in enumerate(sheet_names)
            }
            for future in as_completed(futures):
                try:
                    all_blocks.extend(future.result())
                except Exception as exc:
                    log.warning("XLSX sheet processing error: %s", exc)

        wb.close()
        return all_blocks
