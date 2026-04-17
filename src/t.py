"""
export_quiz_to_excel.py
Đọc file JSON quiz và xuất ra file Excel có định dạng đẹp.

Usage:
    python export_quiz_to_excel.py <input_json> <output_xlsx>
    python export_quiz_to_excel.py quiz_LOMA281.json quiz_LOMA281.xlsx
"""

import json
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Colors ────────────────────────────────────────────────────────────────────
C_HEADER_BG  = "2F5496"   # Dark blue header
C_HEADER_FG  = "FFFFFF"   # White text
C_SUBHDR_BG  = "BDD7EE"   # Light blue sub-header
C_ROW_ODD    = "FFFFFF"
C_ROW_EVEN   = "F2F2F2"
C_CORRECT_BG = "E2EFDA"   # Light green – correct answer column
C_TITLE_BG   = "1F3864"   # Navy title bar
C_BEGINNER   = "70AD47"   # Green
C_INTER      = "FFC000"   # Amber
C_ADVANCED   = "FF0000"   # Red


def thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def header_fill(color):
    return PatternFill("solid", start_color=color, fgColor=color)


def difficulty_color(level: str) -> str:
    level = (level or "").lower()
    if "beginner" in level:
        return C_BEGINNER
    if "intermediate" in level or "medium" in level:
        return C_INTER
    return C_ADVANCED


def build_options_text(options: list) -> str:
    return "\n".join(f"{o['index']}. {o['text']}" for o in options)


def build_source_text(sources: list) -> str:
    parts = []
    for s in sources:
        parts.append(
            f"[{s.get('course','')}] {s.get('name','')} "
            f"– Module {s.get('module','?')} Lesson {s.get('lesson','?')} "
            f"p.{s.get('page','?')}/{s.get('total_pages','?')}"
        )
    return "\n".join(parts)


def export(input_path: str, output_path: str):
    # ── Load JSON ─────────────────────────────────────────────────────────────
    with open(input_path, encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 1 – QUIZ DATA (full detail)
    # ═══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Quiz Data"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Column definitions: (header_text, width, wrap)
    COLS = [
        ("No.",              6,   False),
        ("Course",           10,  False),
        ("Module",           8,   False),
        ("Lesson",           8,   False),
        ("Category",         18,  True),
        ("Node",             18,  True),
        ("Question Type",    22,  True),
        ("Difficulty",       12,  False),
        ("Question",         50,  True),
        ("Option A",         35,  True),
        ("Option B",         35,  True),
        ("Option C",         35,  True),
        ("Option D",         35,  True),
        ("Correct Answer",   14,  False),
        ("Correct Reason",   40,  True),
        ("Hint",             35,  True),
        ("Explanation",      55,  True),
        ("Source",           50,  True),
    ]

    # Title row (row 1)
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    title_cell = ws["A1"]
    title_cell.value = f"Quiz Bank – {os.path.splitext(os.path.basename(input_path))[0]}"
    title_cell.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    title_cell.fill = header_fill(C_TITLE_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Header row (row 2)
    for col_idx, (hdr, width, _wrap) in enumerate(COLS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=hdr)
        cell.font      = Font(name="Arial", bold=True, color=C_HEADER_FG)
        cell.fill      = header_fill(C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 30

    # Data rows (row 3+)
    option_map = {"A": 10, "B": 11, "C": 12, "D": 13}  # col indices for options

    for row_num, q in enumerate(data, start=1):
        r = row_num + 2  # Excel row
        options = q.get("options", [])
        opts_by_idx = {o["index"]: o["text"] for o in options}
        sources = q.get("sources", [{}])
        src = sources[0] if sources else {}

        row_bg = C_ROW_ODD if row_num % 2 == 1 else C_ROW_EVEN

        values = [
            row_num,
            src.get("course", ""),
            src.get("module", ""),
            src.get("lesson", ""),
            q.get("category", ""),
            q.get("node_name", ""),
            q.get("question_type", ""),
            q.get("difficulty", ""),
            q.get("question", ""),
            opts_by_idx.get("A", ""),
            opts_by_idx.get("B", ""),
            opts_by_idx.get("C", ""),
            opts_by_idx.get("D", ""),
            q.get("correct_answer", ""),
            q.get("correct_reason", ""),
            q.get("hint", ""),
            q.get("explanation", ""),
            build_source_text(sources),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col_idx, value=value)
            _wrap = COLS[col_idx - 1][2]
            cell.alignment = Alignment(vertical="top", wrap_text=_wrap)
            cell.border    = thin_border()
            cell.font      = Font(name="Arial", size=10)

            # Correct Answer column → green background
            if col_idx == 14:
                cell.fill = header_fill(C_CORRECT_BG)
                cell.font = Font(name="Arial", size=10, bold=True, color="375623")
                cell.alignment = Alignment(horizontal="center", vertical="top")
            # Difficulty column → colored text
            elif col_idx == 8:
                cell.fill = header_fill(row_bg)
                cell.font = Font(name="Arial", size=10, bold=True,
                                 color=difficulty_color(str(value)))
            else:
                cell.fill = header_fill(row_bg)

        ws.row_dimensions[r].height = 60

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 2 – SUMMARY
    # ═══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False

    # Counts by category
    from collections import Counter
    cat_counter  = Counter(q.get("category", "N/A")     for q in data)
    diff_counter = Counter(q.get("difficulty", "N/A")   for q in data)
    type_counter = Counter(q.get("question_type", "N/A") for q in data)

    def write_summary_section(ws, start_row, title, counter_data):
        ws.merge_cells(f"A{start_row}:C{start_row}")
        hc = ws[f"A{start_row}"]
        hc.value = title
        hc.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        hc.fill  = header_fill(C_HEADER_BG)
        hc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[start_row].height = 22

        sub_row = start_row + 1
        ws.cell(row=sub_row, column=1, value="Name").font  = Font(bold=True, name="Arial")
        ws.cell(row=sub_row, column=2, value="Count").font = Font(bold=True, name="Arial")
        ws.cell(row=sub_row, column=3, value="%").font     = Font(bold=True, name="Arial")
        for c in [1, 2, 3]:
            ws.cell(row=sub_row, column=c).fill   = header_fill(C_SUBHDR_BG)
            ws.cell(row=sub_row, column=c).border = thin_border()

        total = sum(counter_data.values())
        for i, (name, cnt) in enumerate(sorted(counter_data.items()), start=1):
            rr = sub_row + i
            fill_bg = C_ROW_ODD if i % 2 == 1 else C_ROW_EVEN
            ws.cell(row=rr, column=1, value=name).fill   = header_fill(fill_bg)
            ws.cell(row=rr, column=2, value=cnt).fill    = header_fill(fill_bg)
            pct_cell = ws.cell(row=rr, column=3,
                               value=f"=B{rr}/B{sub_row + len(counter_data) + 1}")
            pct_cell.number_format = "0.0%"
            pct_cell.fill = header_fill(fill_bg)
            for c in [1, 2, 3]:
                ws.cell(row=rr, column=c).font   = Font(name="Arial", size=10)
                ws.cell(row=rr, column=c).border = thin_border()
                ws.cell(row=rr, column=c).alignment = Alignment(vertical="center")

        # Total row
        tr = sub_row + len(counter_data) + 1
        ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True, name="Arial")
        ws.cell(row=tr, column=2, value=f"=SUM(B{sub_row+1}:B{tr-1})").font = Font(bold=True, name="Arial")
        ws.cell(row=tr, column=3, value="100%").font = Font(bold=True, name="Arial")
        for c in [1, 2, 3]:
            ws.cell(row=tr, column=c).fill   = header_fill(C_SUBHDR_BG)
            ws.cell(row=tr, column=c).border = thin_border()

        return tr + 2  # next start row

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 10

    next_row = write_summary_section(ws2, 1,  "📁 By Category",      cat_counter)
    next_row = write_summary_section(ws2, next_row, "⚡ By Difficulty", diff_counter)
    write_summary_section(ws2, next_row, "🏷️ By Question Type", type_counter)

    # ── Save ──────────────────────────────────────────────────────────────────
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"✅ Exported {len(data)} questions → {output_path}")


if __name__ == "__main__":
    export(r"/data/quiz_LOMA281_backup.json", r"C:\imt-ai-brain\data\quiz_LOMA281.xlsx")