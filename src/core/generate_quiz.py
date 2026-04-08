"""
Standalone script: Generate quiz using quiz_generator logic and export to Excel.

Usage:
    python generate_quiz_to_excel.py

Parameters (hardcoded – edit below if needed):
    knowledge_pack = "LOMA281"
    quiz_type      = "rate"
    rate_value     = "2|5|3"   (Beginner|Intermediate|Advanced)
    level          = "module"
    level_value    = "2"
    n              = 200
    window         = 2
"""

import asyncio
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ── Ensure project root is on sys.path so `src.*` imports work ────────────────
PROJECT_ROOT = Path(__file__).resolve().parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.core.quiz.quiz_generator import generate_quiz  # noqa: E402

# ═══════════════════════════════════════════════════════════════════════════════
# PARAMETERS – edit here
# ═══════════════════════════════════════════════════════════════════════════════
KNOWLEDGE_PACK = "LOMA281"
QUIZ_TYPE = "rate"
RATE_VALUE = "2|5|3"
LEVEL = "module"
LEVEL_VALUE = "2"
N = 200
WINDOW = 2

OUTPUT_DIR = PROJECT_ROOT / "data" / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL WRITER
# ═══════════════════════════════════════════════════════════════════════════════

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
META_FONT = Font(bold=True, size=11)
META_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def _apply_style(ws, row, col, font=None, fill=None):
    cell = ws.cell(row=row, column=col)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    cell.border = THIN_BORDER
    cell.alignment = WRAP_ALIGNMENT


def _extract_wrong_reasons(question: dict) -> dict[int, str]:
    """Return {index: reason} for wrong options from traps.wrong_reasons."""
    traps = question.get("traps") or {}
    wrong_reasons_list = traps.get("wrong_reasons") or []
    return {wr.get("index"): wr.get("reason", "") for wr in wrong_reasons_list}


def write_excel(result: dict, output_path: Path):
    """Write quiz result to an Excel file."""
    data = result.get("data", result)
    questions = data.get("questions", [])

    total = data.get("total", len(questions))
    knowledge_pack = data.get("knowledge_pack", "")
    level = data.get("level")
    level_value = data.get("level_value")
    quiz_type = data.get("type", "")
    dist = data.get("difficulty_distribution", {})
    dist_counts = dist.get("counts", {})
    dist_rate = dist.get("rate", {})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quiz"

    # ── Row 1-8: Metadata ─────────────────────────────────────────────────────
    meta_rows = [
        ("Total", total),
        ("Knowledge Pack", knowledge_pack),
        ("Level", level),
        ("Level Value", level_value),
        ("Type", quiz_type),
        ("Beginner Count", dist_counts.get("Beginner", 0)),
        ("Intermediate Count", dist_counts.get("Intermediate", 0)),
        ("Advanced Count", dist_counts.get("Advanced", 0)),
        ("Beginner Rate (%)", dist_rate.get("Beginner", 0)),
        ("Intermediate Rate (%)", dist_rate.get("Intermediate", 0)),
        ("Advanced Rate (%)", dist_rate.get("Advanced", 0)),
    ]

    for i, (label, value) in enumerate(meta_rows, start=1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
        _apply_style(ws, i, 1, font=META_FONT, fill=META_FILL)
        _apply_style(ws, i, 2)

    # ── Header row for questions ──────────────────────────────────────────────
    HEADER_ROW = len(meta_rows) + 2  # blank row then header

    columns = [
        "Question",
        "Answer 1",
        "Answer 2",
        "Answer 3",
        "Answer 4",
        "Correct Answer",
        "Difficulty Level",
        "Hint",
        "Source File Name",
        "Source Page",
        "Source Module",
        "Source Lesson",
        "Source Course",
        "Explanation",
        "Category",
        "Node Name",
        "Trap Description",
        "Reason Wrong Answer 1",
        "Reason Wrong Answer 2",
        "Reason Wrong Answer 3",
        "Reason Correct Answer",
        "Question Type",
    ]

    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=HEADER_ROW, column=col_idx, value=col_name)
        _apply_style(ws, HEADER_ROW, col_idx, font=HEADER_FONT, fill=HEADER_FILL)

    # ── Data rows ─────────────────────────────────────────────────────────────
    for q_idx, q in enumerate(questions, start=1):
        row = HEADER_ROW + q_idx

        # Options
        options = q.get("options") or []
        option_texts = [""] * 4
        for opt in options:
            idx = opt.get("index", 0)
            if 0 <= idx < 4:
                option_texts[idx] = opt.get("text", "")

        # Correct answer text
        correct_idx = q.get("correct_answer", 0)
        correct_text = option_texts[correct_idx] if 0 <= correct_idx < 4 else ""

        # Sources (take first source)
        sources = q.get("sources") or []
        src = sources[0] if sources else {}
        source_file = src.get("name", "")
        source_page = src.get("page", "")
        source_module = src.get("module", "")
        source_lesson = src.get("lesson", "")
        source_course = src.get("course", "")

        # Traps
        traps = q.get("traps") or {}
        trap_desc = traps.get("description", "")
        wrong_reasons = _extract_wrong_reasons(q)

        # Build wrong reason columns: for non-correct options in order
        wrong_option_indices = [i for i in range(4) if i != correct_idx]
        reason_wrong = [wrong_reasons.get(i, "") for i in wrong_option_indices]
        while len(reason_wrong) < 3:
            reason_wrong.append("")

        correct_reason = q.get("correct_reason", "")

        values = [
            q.get("question", ""),
            option_texts[0],
            option_texts[1],
            option_texts[2],
            option_texts[3],
            correct_text,
            q.get("difficulty", ""),
            q.get("hint", ""),
            source_file,
            source_page,
            source_module,
            source_lesson,
            source_course,
            q.get("explanation", ""),
            q.get("category", ""),
            q.get("node_name", ""),
            trap_desc,
            reason_wrong[0],
            reason_wrong[1],
            reason_wrong[2],
            correct_reason,
            q.get("question_type", ""),
        ]

        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=row, column=col_idx, value=val)
            _apply_style(ws, row, col_idx)

    # ── Auto-width columns (capped at 60) ────────────────────────────────────
    for col_idx in range(1, len(columns) + 1):
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        for cell in ws[col_letter]:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[col_letter].width = max_len + 4

    wb.save(output_path)
    print(f"\n✅ Excel saved → {output_path}")
    print(f"   Total questions: {total}")
    print(f"   Difficulty: {dist_counts}")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

async def main():
    print("=" * 70)
    print("  Quiz Generator → Excel")
    print("=" * 70)
    print(f"  Knowledge Pack : {KNOWLEDGE_PACK}")
    print(f"  Type           : {QUIZ_TYPE}")
    print(f"  Rate Value     : {RATE_VALUE}")
    print(f"  Level          : {LEVEL}")
    print(f"  Level Value    : {LEVEL_VALUE}")
    print(f"  N              : {N}")
    print(f"  Window         : {WINDOW}")
    print("=" * 70)

    print("\n⏳ Generating quiz (this may take a few minutes for n=200)...\n")

    result = await generate_quiz(
        knowledge_pack=KNOWLEDGE_PACK,
        level=LEVEL,
        level_value=LEVEL_VALUE,
        quiz_type=QUIZ_TYPE,
        rate_value=RATE_VALUE,
        n=N,
        window=WINDOW,
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"quiz_{KNOWLEDGE_PACK}_module{LEVEL_VALUE}_{N}q_{timestamp}.xlsx"
    output_path = OUTPUT_DIR / filename

    write_excel(result, output_path)


if __name__ == "__main__":
    asyncio.run(main())
